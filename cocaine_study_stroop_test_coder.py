from psychopy import core, visual, gui, data, event
from psychopy.tools.filetools import fromFile, toFile
import numpy, random, csv, json, os, openpyxl
from openpyxl import Workbook


expInfo = {'Subject ID': '', 'Session Number': '', 'Start Program From': ['Beginning (Loop 1)','Loop 2','Loop 3']} ###Change to subject id
expInfo['dateStr'] = data.getDateStr()

########present a dialogue to change params

dlg = gui.DlgFromDict(expInfo, title='Cocaine Study Stroop Test', fixed=['dateStr'], order=['Subject ID','Session Number','Start Program From'])
#dlg.addField(initial = 'Start Program From:', choices=['Beginning (Loop 1)','Loop 2','Loop 3'])
if dlg.OK:
    toFile('lastParams.pickle', expInfo) #save params to file for next time
else:
    core.quit() #the user hit cancel, so exit
    
#########Definition of Variable start_loop --> Used to determine which loop the program undergoes

start_loop = ''

if expInfo['Start Program From'] == 'Beginning (Loop 1)':
    start_loop = 'loop_1'
elif expInfo['Start Program From'] == 'Loop 2':
    start_loop = 'loop_2'
elif expInfo['Start Program From'] == 'Loop 3':
    start_loop = 'loop_3'

print('Starting Program with ' + start_loop)

#########Definition of Global Clock --> Used to keep track of the time words/fixation crosses appear and disappear, and GLOBAL time the first button was clicked

globalClock = core.Clock()


############make a csv file to store the data- intro data
fileName_intro_data = expInfo['Subject ID'] + '_' + expInfo['Session Number'] + '_start_from_' + start_loop +'_IntroData_CocaineStroopTest' + expInfo['dateStr']
dataFile_intro_data = open('data/' + fileName_intro_data + '.csv', 'w') # a simple text file with comma seperated values
#dataFile.write('sequence,thisN,thisRepN,word,wordtype,number_on_screen,correctAnswer,key_pressed,correct,time_button_pressed_relative, buttonClickedList, time_word_shown_global, time_button_pressed_global, time_word_gone_global, time_cross_shown_global, time_cross_gone_global \n')

############make a csv file to store the data- test data
fileName = expInfo['Subject ID'] + '_' + expInfo['Session Number'] + '_start_from_' + start_loop + '_CocaineStroopTest' + expInfo['dateStr']
dataFile = open('data/' + fileName + '.csv', 'w') # a simple text file with comma seperated values
dataFile.write('sequence,thisN,thisRepN,word,wordtype,number_on_screen,correctAnswer,key_pressed,correct,time_button_pressed_relative, buttonClickedList, time_word_shown_global, time_button_pressed_global, time_word_gone_global, time_cross_shown_global, time_cross_gone_global \n')



#############import main.xlsx
mainlist = data.importConditions('main.xlsx')

#############get pre-seq prior from mainlist (needs to be pre because personal words will be inserted later once personal word number is determined to either be set or random)
preseq_1_1 = mainlist[0]
preseq_1_2_ic = mainlist[1]
preseq_1_3 = mainlist[2]

preseq_2_1 = mainlist[3]
preseq_2_2_ic = mainlist[4]
preseq_2_3 = mainlist[5]

preseq_3_1 = mainlist[6]
preseq_3_2_ic = mainlist[7]
preseq_3_3 = mainlist[8]

####################################################################################################
'''
Step 1. Load personal_words_randomization.xlsx.
    - if personal_word_randomization is "none", then proceed
Step 2. Load list_of_twelve_personal_words.xlsx twice into a 2d list. list[0] contains the list random


Everything below this line is no longer necessary because randomization only occurs once per participant, the number of times the personalized words appear in
a frame is fixed, personal word slots are fixed, and the list of 12 words must be randomized, then inserted into slots, then 12 words are randomized again, and
then inserted into remaining slots.
'''
###################################################################################################
###Personal Word Randomizer Function -> If personal_words/personal_words_order_and_number.xlsx file does not exist, make one. It is created by  
###taking the list_of_twelve_personal_words.xlsx file (which is the 12 personal words collected from screening and inserted in folder before running the program)
###and using it to create a set of 24 words that are appeared. The new file is a unique, random location of the words

def personal_word_randomizer():
    personal_word_list = data.importConditions('personal_words/list_of_eight_personal_words.xlsx')
    
    larger_list = []
    amount = len(personal_word_list)
    for x in range(8):
        random_index = random.randrange(0,len(personal_word_list))
        larger_list.append(personal_word_list[random_index]['eight_words'])
        personal_word_list.pop(random_index)
    
    second_personal_word_list = data.importConditions('personal_words/list_of_eight_personal_words.xlsx')
    second_amount = len(second_personal_word_list)
    for x in range(8):
        random_index = random.randrange(0,len(second_personal_word_list))
        larger_list.append(second_personal_word_list[random_index]['eight_words'])
        second_personal_word_list.pop(random_index)
        
    third_personal_word_list = data.importConditions('personal_words/list_of_eight_personal_words.xlsx')
    third_amount = len(third_personal_word_list)
    for x in range(8):
        random_index = random.randrange(0,len(third_personal_word_list))
        larger_list.append(third_personal_word_list[random_index]['eight_words'])
        third_personal_word_list.pop(random_index)
        
    number_answer_list = [[1,2],[1,2],[4,5],[1,2],[3,4],[2,3],[4,5],[3,4],[2,3],[3,4],[2,3],[4,5],[1,2],[1,2],[3,4],[2,3],[3,4],[2,3],[2,3],[4,5],[4,5],[3,4],[4,5],[1,2]]
    
    personal_word_dataFile = openpyxl.Workbook() #Create a new excel sheet to store the randomized words
    ws = personal_word_dataFile.active
    ws['A1'] = 'personal_words_order'
    ws['B1'] = 'type'
    ws['C1'] = 'number'
    ws['D1'] = 'answer'
    
    personal_word_type = "personal"
    
    rows = []
    for x in range(len(larger_list)):
        rows.append((larger_list[x],personal_word_type,number_answer_list[x][0],number_answer_list[x][1]))
        
    for row in rows:
        ws.append(row)
        personal_word_dataFile.save('personal_words/personal_words_order_and_number.xlsx') #name the excel sheet 
        
    
#########Check if personal_word_order_and_number exists.xlsx exists. If it does not, then create a new excel sheet
personal_word_order_and_number_file_path = 'personal_words/personal_words_order_and_number.xlsx'

if os.path.isfile(personal_word_order_and_number_file_path):
    print('personal_words_order_and_number.xlsx already exists in folder. Proceeding to program.')
    personal_words = data.importConditions('personal_words/personal_words_order_and_number.xlsx')
else:
    print('personal_words_order_and_number.xlsx does not exist in folder. Creating one now...')
    personal_word_randomizer()
    personal_words = data.importConditions('personal_words/personal_words_order_and_number.xlsx')
    
    
############
def personal_word_inserter(preseq):
    preseqlist = data.importConditions(preseq["blocks"])
    for i in range(0,len(preseqlist)):
        if preseqlist[i]["word"] == None:
            #preseqlist[i] = personal_words.pop(0)
            newitem = personal_words.pop(0)
            preseqlist[i]["word"] = newitem["personal_words_order"]
            preseqlist[i]["type"] = newitem["type"]
            preseqlist[i]["number"] = newitem["number"]
            preseqlist[i]["answer"] = newitem["answer"]
    #if preseq["personal_word_number"] == "random":  ##--> Not necessary right now
        #personal_word_number_randomizer(preseqlist)  ##--> Not necessary right now
    return preseqlist 
    
seq_1_1 = data.TrialHandler(trialList=personal_word_inserter(preseq_1_1),nReps=1,method='sequential',originPath=None)
seq_1_2_ic = data.TrialHandler(trialList=personal_word_inserter(preseq_1_2_ic),nReps=1,method='sequential',originPath=None)
seq_1_3 = data.TrialHandler(trialList=personal_word_inserter(preseq_1_3),nReps=1,method='sequential',originPath=None)
    
seq_2_1 = data.TrialHandler(trialList=personal_word_inserter(preseq_2_1),nReps=1,method='sequential',originPath=None)
seq_2_2_ic = data.TrialHandler(trialList=personal_word_inserter(preseq_2_2_ic),nReps=1,method='sequential',originPath=None)
seq_2_3 = data.TrialHandler(trialList=personal_word_inserter(preseq_2_3),nReps=1,method='sequential',originPath=None)

seq_3_1 = data.TrialHandler(trialList=personal_word_inserter(preseq_3_1),nReps=1,method='sequential',originPath=None)
seq_3_2_ic = data.TrialHandler(trialList=personal_word_inserter(preseq_3_2_ic),nReps=1,method='sequential',originPath=None)
seq_3_3 = data.TrialHandler(trialList=personal_word_inserter(preseq_3_3),nReps=1,method='sequential',originPath=None)


win = visual.Window(fullscr=True,allowGUI=True, checkTiming=True)
#win = visual.Window([800,800])
#event.globalKeys.add(key=quitKey, func=forceQuit)
welcome_message = visual.TextStim(win, pos=[0,0], text='Welcome to the Stroop Test!', height=0.20)
fixation_cross = visual.TextStim(win, text="+", height=1)
instruction_1_text = 'Count the number of words you see on the screen.' + '\n\n' + 'Press the button that represents the amount of words you counted as fast as you can.' + '\n\n' + 'Let\'s practice! Press BUTTON 1 (index finger) now'
instruction_2_text = 'Count the number of words you see on the screen.' + '\n\n' + 'Press the button that represents the amount of words you counted as fast as you can.' + '\n\n' + 'Let\'s practice! Press BUTTON 2 (middle finger) now'
instruction_3_text = 'Count the number of words you see on the screen.' + '\n\n' + 'Press the button that represents the amount of words you counted as fast as you can.' + '\n\n' + 'Let\'s practice! Press BUTTON 3 (ring finger) now'
instruction_4_text = 'Count the number of words you see on the screen.' + '\n\n' + 'Press the button that represents the amount of words you counted as fast as you can.' + '\n\n' + 'Let\'s practice! Press BUTTON 4 (small finger/pinky) now'
start_screen_text = 'Great! Now let\'s start'
get_ready_text = 'Get Ready'

instruction_1_message = visual.TextStim(win, pos=[0,0], text=instruction_1_text, height=0.13)
instruction_2_message = visual.TextStim(win, pos=[0,0], text=instruction_2_text, height=0.13)
instruction_3_message = visual.TextStim(win, pos=[0,0], text=instruction_3_text, height=0.13)
instruction_4_message = visual.TextStim(win, pos=[0,0], text=instruction_4_text, height=0.13)

start_screen_message = visual.TextStim(win, pos=[0,0], text=start_screen_text, height=0.20)
get_ready_message = visual.TextStim(win, pos=[0,0], text=get_ready_text, height=0.20)
blank_infinite = visual.TextStim(win, pos=[0,0], text="")
goodbye_message = visual.TextStim(win, pos=[0,0], text="Thanks for participating!", height=0.20)

#blocks,sequence,thisN,thisRepN,word,type,number,answer,key_pressed,correct,time_button_pressed,time_fixation_cross_appeared,time_fixation_cross_stopped,duration_fixation_cross,time_word_appeared,time_word_stopped,duration_word
    
#Definition of Repeating Words Function
def newWordText(increment):
    counter = int(increment["number"])
    newText = ""
    while counter > 0:
        newText += increment["word"] + "\n"
        counter -= 1
    return newText
  
#Definition of Trial Clock --> Will reset everytime a new word is shown in order to get relative time the first button was clicked
trialClock = core.Clock()

#Definition of Loop Function
def Loop(first_seq, first_seqname, second_seq, second_seqname, third_seq, third_seqname):
    
    
    ####First Sequence Loop###############################################################################################################################
    new_img = visual.TextStim(win, pos=[0,0], height=0.2, bold=True)
    for thisIncrement in first_seq:
        
        ##Quit Button During Code
        
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
        
        buttonsClickedList = []
        displaytext = newWordText(thisIncrement)
        
        ##Setting important variables for this specific word
        new_img.text = displaytext
        word = thisIncrement["word"]
        wordtype = thisIncrement["type"]
        number_on_screen = thisIncrement["number"]
        correctAnswer = thisIncrement["answer"]
        thisN = first_seq.thisN
        thisRepN = first_seq.thisRepN

        key_pressed = None
        correct = None
        sequence = first_seqname
        time_button_pressed_relative = None
        
        time_word_shown_global = None
        time_button_pressed_global = None
        time_word_gone_global = None
        time_cross_shown_global = None
        time_cross_gone_global = None
        
       
        time_button_pressed_global_bool = True ##Exists to make sure that the global time for only the first button press is recorded, not all the others
        
        
        
        ##Reset trialClock Right before image is shown for 2 seconds
        trialClock.reset()

        ##The 2 seconds (120 frames) where the image is shown
        time_word_shown_global = str(globalClock.getTime())
        for x in range(120):
            
    
            ##Quit Button During Code
            quitbutton = event.getKeys(keyList=['escape'])
            if len(quitbutton) > 0:
                for a in quitbutton:
                    if a == 'escape':
                        quitTime = str(globalClock.getTime())
                        dataFile.write(f"\nQuit_Time,{quitTime}")
                        dataFile.close()
                        dataFile_intro_data.close()
                        core.quit()
            
            
            #new_img.text = displaytext 
            new_img.draw() 
            win.flip() 
            allKeys = event.getKeys(keyList=['2','3','4','5'], timeStamped=True) 
            if len(allKeys) > 0: 
                time_button_pressed_relative = str(trialClock.getTime()) 
                allKeys[0][1] = time_button_pressed_relative  
                buttonsClickedList.append(allKeys[0]) 
                
                if time_button_pressed_global_bool == True:
                    time_button_pressed_global = str(globalClock.getTime())
                    time_button_pressed_global_bool = False
        
        time_word_gone_global = str(globalClock.getTime())
        trialClock.reset() 
                
        if len(buttonsClickedList) > 0:
            key_pressed = buttonsClickedList[0][0]
            time_button_pressed_relative = buttonsClickedList[0][1]
        if key_pressed == None:
            correct = None
        elif int(key_pressed) == int(correctAnswer):
            correct = True
        else:
            correct = False
        
        #json_buttonsClickedList = json.dumps(buttonsClickedList)
        flat = '; '.join([':: '.join(sublist) for sublist in buttonsClickedList])
        
        #Write data to csv file up to time_word_gone_global
        dataFile.write(f"{sequence},{thisN},{thisRepN},{word},{wordtype},{number_on_screen},{correctAnswer},{key_pressed},{correct},{time_button_pressed_relative},{flat},{time_word_shown_global},{time_button_pressed_global},{time_word_gone_global},")
        
        ##300 ms fixation cross
        time_cross_shown_global = str(globalClock.getTime())
        for x in range(18):
            
            
            ##Quit Button During Code
            quitbutton = event.getKeys(keyList=['escape'])
            if len(quitbutton) > 0:
                for a in quitbutton:
                    if a == 'escape':
                        quitTime = str(globalClock.getTime())
                        dataFile.write(f"\nQuit_Time,{quitTime}")
                        dataFile.close()
                        dataFile_intro_data.close()
                        core.quit()
            
            
            fixation_cross.draw()
            win.flip()
            #Remember to record appearance of fixation cross ##Actually, maybe not...
        time_cross_gone_global = str(globalClock.getTime())
        
        #Write data to csv for time_cross_shown_global and time_cross_gone_global
        dataFile.write(f"{time_cross_shown_global},{time_cross_gone_global}\n")
        
    
    
    
    ##Interim Fixation Between First and Second Loop#####################################################################################################
    for x in range(1200):
        fixation_cross.draw()
        win.flip()
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        #Remember to record appearance of fixation cross ##Actually, maybe not...
    
    ####Second Sequence Loop###############################################################################################################################
    new_img = visual.TextStim(win, pos=[0,0], height=0.2, bold=True)
    for thisIncrement in second_seq:
        
        
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
        
        buttonsClickedList = []
        displaytext = newWordText(thisIncrement)
        
        ##Setting important variables for this specific word
        new_img.text = displaytext
        word = thisIncrement["word"]
        wordtype = thisIncrement["type"]
        number_on_screen = thisIncrement["number"]
        correctAnswer = thisIncrement["answer"]
        thisN = second_seq.thisN
        thisRepN = second_seq.thisRepN
        key_pressed = None
        correct = None
        sequence = second_seqname
        time_button_pressed_relative = None
        
        time_word_shown_global = None
        time_button_pressed_global = None
        time_word_gone_global = None
        time_cross_shown_global = None
        time_cross_gone_global = None 
        
        time_button_pressed_global_bool = True ##Exists to make sure that the global time for only the first button press is recorded, not all the others
        
        ##Reset trialClock Right before image is shown for 2 seconds
        trialClock.reset()

        ##The 2 seconds (120 frames) where the image is shown
        time_word_shown_global = str(globalClock.getTime())
        for x in range(120): 
            
            ##Quit Button During Code
            quitbutton = event.getKeys(keyList=['escape'])
            if len(quitbutton) > 0:
                for a in quitbutton:
                    if a == 'escape':
                        quitTime = str(globalClock.getTime())
                        dataFile.write(f"\nQuit_Time,{quitTime}")
                        dataFile.close()
                        dataFile_intro_data.close()
                        core.quit()
            
            #new_img.text = displaytext 
            new_img.draw() 
            win.flip() 
            allKeys = event.getKeys(keyList=['2','3','4','5'], timeStamped=True) 
            if len(allKeys) > 0: 
                time_button_pressed_relative = str(trialClock.getTime()) 
                allKeys[0][1] = time_button_pressed_relative  
                buttonsClickedList.append(allKeys[0]) 
                
                if time_button_pressed_global_bool == True:
                    time_button_pressed_global = str(globalClock.getTime())
                    time_button_pressed_global_bool = False
        
        time_word_gone_global = str(globalClock.getTime())
        trialClock.reset() 
                
        
        if len(buttonsClickedList) > 0:
            key_pressed = buttonsClickedList[0][0]
            time_button_pressed_relative = buttonsClickedList[0][1]
        if key_pressed == None:
            correct = None
        elif int(key_pressed) == int(correctAnswer):
            correct = True
        else:
            correct = False
            
        #json_buttonsClickedList = json.dumps(buttonsClickedList)
        flat = '; '.join([':: '.join(sublist) for sublist in buttonsClickedList])
        
        dataFile.write(f"{sequence},{thisN},{thisRepN},{word},{wordtype},{number_on_screen},{correctAnswer},{key_pressed},{correct},{time_button_pressed_relative},{flat},{time_word_shown_global}, {time_button_pressed_global},{time_word_gone_global},")
        
        ##300 ms fixation cross
        time_cross_shown_global = str(globalClock.getTime())
        for x in range(18):
            
            
            ##Quit Button During Code
            quitbutton = event.getKeys(keyList=['escape'])
            if len(quitbutton) > 0:
                for a in quitbutton:
                    if a == 'escape':
                        quitTime = str(globalClock.getTime())
                        dataFile.write(f"\nQuit_Time,{quitTime}")
                        dataFile.close()
                        dataFile_intro_data.close()
                        core.quit()
            
            
            fixation_cross.draw()
            win.flip()
            #Remember to record appearance of fixation cross ##Actually, maybe not...
        
        time_cross_gone_global = str(globalClock.getTime())
        
        #Write data to csv for time_cross_shown_global and time_cross_gone_global
        dataFile.write(f"{time_cross_shown_global},{time_cross_gone_global}\n")
        
            
    ##Interim Fixation Between Second and Third Loop#######################################################################################################
    for x in range(1200):
        fixation_cross.draw()
        win.flip()
        #Remember to record appearance of fixation cross ##Actually, maybe not...
        
        
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
    
    ####Third Sequence Loop################################################################################################################################
    new_img = visual.TextStim(win, pos=[0,0], height=0.2, bold=True)
    for thisIncrement in third_seq:
        
        
        
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
        
        buttonsClickedList = []
        displaytext = newWordText(thisIncrement)
        
        ##Setting important variables for this specific word
        new_img.text = displaytext
        word = thisIncrement["word"]
        wordtype = thisIncrement["type"]
        number_on_screen = thisIncrement["number"]
        correctAnswer = thisIncrement["answer"]
        thisN = third_seq.thisN
        thisRepN = third_seq.thisRepN
        key_pressed = None
        correct = None
        sequence = third_seqname
        time_button_pressed_relative = None
        
        time_word_shown_global = None
        time_button_pressed_global = None
        time_word_gone_global = None
        time_cross_shown_global = None
        time_cross_gone_global = None 
        
        time_button_pressed_global_bool = True ##Exists to make sure that the global time for only the first button press is recorded, not all the others
            
        ##Reset trialClock Right before image is shown for 2 seconds
        trialClock.reset()
            
        ##The 2 seconds (120 frames) where the image is shown
        time_word_shown_global = str(globalClock.getTime())
        for x in range(120): 
            
            
            ##Quit Button During Code
            quitbutton = event.getKeys(keyList=['escape'])
            if len(quitbutton) > 0:
                for a in quitbutton:
                    if a == 'escape':
                        quitTime = str(globalClock.getTime())
                        dataFile.write(f"\nQuit_Time,{quitTime}")
                        dataFile.close()
                        dataFile_intro_data.close()
                        core.quit()
            
            
            #new_img.text = displaytext 
            new_img.draw() 
            win.flip() 
            allKeys = event.getKeys(keyList=['2','3','4','5'], timeStamped=True) 
            if len(allKeys) > 0: 
                time_button_pressed_relative = str(trialClock.getTime()) 
                allKeys[0][1] = time_button_pressed_relative  
                buttonsClickedList.append(allKeys[0]) 
                
                if time_button_pressed_global_bool == True:
                    time_button_pressed_global = str(globalClock.getTime())
                    time_button_pressed_global_bool = False
            
        time_word_gone_global = str(globalClock.getTime())
        trialClock.reset() 
                
            
        if len(buttonsClickedList) > 0:
            key_pressed = buttonsClickedList[0][0]
            time_button_pressed_relative = buttonsClickedList[0][1]
        if key_pressed == None:
            correct = None
        elif int(key_pressed) == int(correctAnswer):
            correct = True
        else:
            correct = False
            
        #json_buttonsClickedList = json.dumps(buttonsClickedList)
        flat = '; '.join([':: '.join(sublist) for sublist in buttonsClickedList])
        
        dataFile.write(f"{sequence},{thisN},{thisRepN},{word},{wordtype},{number_on_screen},{correctAnswer},{key_pressed},{correct},{time_button_pressed_relative},{flat},{time_word_shown_global}, {time_button_pressed_global},{time_word_gone_global},")
        
        ##300 ms fixation cross
        time_cross_shown_global = str(globalClock.getTime())
        for x in range(18):
            
            
            ##Quit Button During Code
            quitbutton = event.getKeys(keyList=['escape'])
            if len(quitbutton) > 0:
                for a in quitbutton:
                    if a == 'escape':
                        quitTime = str(globalClock.getTime())
                        dataFile.write(f"\nQuit_Time,{quitTime}")
                        dataFile.close()
                        dataFile_intro_data.close()
                        core.quit()
            
            
            fixation_cross.draw()
            win.flip()
            #Remember to record appearance of fixation cross ##Actually, maybe not...
            
        time_cross_gone_global = str(globalClock.getTime())
        #Write data to csv for time_cross_shown_global and time_cross_gone_global
        dataFile.write(f"{time_cross_shown_global},{time_cross_gone_global}\n")
    
        
#0. Establish variables for instructions stuff
message_number = ""
message_content = ""
time_message_shown = ""
time_message_gone = ""
button_pressed = ""
time_button_pressed = ""
empty_cell = ""
cross_type = ""
time_cross_shown = ""
time_cross_gone = ""


#1. Display Welcome Screen, infinite until "t" is pressed. Also create press_t eventKeys object
'#welcome message shown'
welcome_message.draw()
message_number = "message0"
message_content = "Welcome to the Stroop Test! Press t to continue."
time_message_shown = str(globalClock.getTime())
win.flip()

'#welcome message gone'
press_t = event.waitKeys(keyList="t", timeStamped=True)
button_pressed = press_t[0][0]
time_button_pressed = str(globalClock.getTime())
dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell},time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


#2. Display initial 300ms (18 frames) Cross Fixation
'#first cross shown'
cross_type = "Cross_After_Welcome_Message"
time_cross_shown = str(globalClock.getTime())
for x in range(18):
    fixation_cross.draw()
    win.flip()
    ##Quit Button During Code
    quitbutton = event.getKeys(keyList=['escape'])
    if len(quitbutton) > 0:
        for a in quitbutton:
            if a == 'escape':
                quitTime = str(globalClock.getTime())
                dataFile.write(f"\nQuit_Time,{quitTime}")
                dataFile.close()
                dataFile_intro_data.close()
                core.quit()
'#first cross gone'
time_cross_gone = str(globalClock.getTime())
dataFile_intro_data.write(f"cross_type,{cross_type},time_cross_shown,{time_cross_shown},{empty_cell},time_cross_gone,{time_cross_gone}\n\n")
    
#3. Instruction 1
'#Instruction 1 shown'
instruction_1_message.draw()
win.flip()
message_number = 'message1'
message_content = 'In this task you will count the number of words you see on the screen. Then press the button as fast as you can to indicate the number of words you counted. Let\'s practice! Press BUTTON 1 (index finger) now'
time_message_shown = str(globalClock.getTime())

'#Instruction 1 gone'
press_2 = event.waitKeys(keyList=["2",'escape'], timeStamped=True)
if press_2[0][0] == 'escape':
    quitTime = str(globalClock.getTime())
    dataFile.write(f"\nQuit_Time,{quitTime}")
    dataFile.close()
    dataFile_intro_data.close()
    core.quit()
button_pressed = press_2[0][0]
time_button_pressed = str(globalClock.getTime())
dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell},time_button_pressed,{time_button_pressed},{empty_cell}\n\n")

#4. Instruction 2
'#Instruction 2 Shown'
instruction_2_message.draw()
win.flip()
message_number = 'message2'
message_content = 'In this task you will count the number of words you see on the screen. Then press the button as fast as you can to indicate the number of words you counted. Let\'s practice! Press BUTTON 2 (middle finger) now'
time_message_shown = str(globalClock.getTime())

'#Instruction 2 gone'
press_3 = event.waitKeys(keyList=["3",'escape'], timeStamped=True)
if press_3[0][0] == 'escape':
    quitTime = str(globalClock.getTime())
    dataFile.write(f"\nQuit_Time,{quitTime}")
    dataFile.close()
    dataFile_intro_data.close()
    core.quit()
button_pressed = press_3[0][0]
time_button_pressed = str(globalClock.getTime())
dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell},time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


#5. Instruction 3
'#Instruction 3 shown'
instruction_3_message.draw()
win.flip()
message_number = 'message3'
message_content = 'In this task you will count the number of words you see on the screen. Then press the button as fast as you can to indicate the number of words you counted. Let\'s practice! Press BUTTON 3 (index finger) now'
time_message_shown = str(globalClock.getTime())

'#Instruction 3 Gone'
press_4 = event.waitKeys(keyList=["4", 'escape'],timeStamped=True)
if press_4[0][0] == 'escape':
    quitTime = str(globalClock.getTime())
    dataFile.write(f"\nQuit_Time,{quitTime}")
    dataFile.close()
    dataFile_intro_data.close()
    core.quit()
button_pressed = press_4[0][0]
time_button_pressed = str(globalClock.getTime())
dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell},time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


#6. Instruction 4
'#Instruction 4 Shown'
instruction_4_message.draw()
win.flip()
message_number = 'message4'
message_content = 'In this task you will count the number of words you see on the screen. Then press the button as fast as you can to indicate the number of words you counted. Let\'s practice! Press BUTTON 4 (small finger/pinky) now'
time_message_shown = str(globalClock.getTime())


'#Instruction 4 Gone'
press_5 = event.waitKeys(keyList=["5", 'escape'], timeStamped=True)
if press_5[0][0] == 'escape':
    quitTime = str(globalClock.getTime())
    dataFile.write(f"\nQuit_Time,{quitTime}")
    dataFile.close()
    dataFile_intro_data.close()
    core.quit()
button_pressed = press_5[0][0]
time_button_pressed = str(globalClock.getTime())
dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell},time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


#7. Display start screen for 5 seconds (300 frames)

message_number = "message5"
message_content = 'Great! Now let\'s start'
time_message_shown = str(globalClock.getTime())
for x in range(300):
    
    ##Quit Button During Code
    quitbutton = event.getKeys(keyList=['escape'])
    if len(quitbutton) > 0:
        for a in quitbutton:
            if a == 'escape':
                quitTime = str(globalClock.getTime())
                dataFile.write(f"\nQuit_Time,{quitTime}")
                dataFile.close()
                dataFile_intro_data.close()
                core.quit()
    
    start_screen_message.draw()
    win.flip()
time_message_gone = str(globalClock.getTime())
dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},time_message_gone,{time_message_gone},{empty_cell}\n\n")


#################################-Loop 1 Function

def runLoop1():
    #8. Display get_ready screen for 10 seconds (600 frames)
    message_number = "Pre_Loop_1_Get_Ready"
    message_content = 'Get Ready'
    time_message_shown = str(globalClock.getTime())
    for x in range(600):
        
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
        get_ready_message.draw()
        win.flip()
    time_message_gone = str(globalClock.getTime())
    dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},time_message_gone,{time_message_gone},{empty_cell}\n\n")


    #9. Display 300ms (18 frames) Cross Fixation Prior to Loop 1
    cross_type = "Pre_Loop_1_Cross"
    time_cross_shown = str(globalClock.getTime())
    for x in range(18):
        
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
        fixation_cross.draw()
        win.flip()
    time_cross_gone = str(globalClock.getTime())
    dataFile_intro_data.write(f"cross_type,{cross_type},time_cross_shown,{time_cross_shown},{empty_cell},time_cross_gone,{time_cross_gone}\n\n")

    #10 Loop 1

    Loop(seq_1_1, "seq_1_1", seq_1_2_ic, "seq_1_2_ic", seq_1_3, "seq_1_3")


    #11 Display Blank Infinite after Loop 1 (press b to continue)
    message_number = 'After_Loop_1_Infinite_Blank'
    message_content = 'Blank screen (must press b to continue)'
    blank_infinite.draw()
    win.flip()
    time_message_shown = str(globalClock.getTime())
    press_b = event.waitKeys(keyList=["b", 'escape'], timeStamped=True)
    if press_b[0][0] == 'escape':
        quitTime = str(globalClock.getTime())
        dataFile.write(f"\nQuit_Time,{quitTime}")
        dataFile.close()
        dataFile_intro_data.close()
        core.quit()
    button_pressed = press_b[0][0]
    time_button_pressed = str(globalClock.getTime())
    dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell}, time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


#####################################-Loop 2 Function

def runLoop2():
    #12 Display Get Ready (press t to continue)
    message_number = "Pre_Loop_2_Get_Ready"
    message_content = "Get Ready"
    get_ready_message.draw()
    win.flip()
    time_message_shown = str(globalClock.getTime())
    press_t_2 = event.waitKeys(keyList=["t", 'escape'], timeStamped=True)
    if press_t_2[0][0] == 'escape':
        quitTime = str(globalClock.getTime())
        dataFile.write(f"\nQuit_Time,{quitTime}")
        dataFile.close()
        dataFile_intro_data.close()
        core.quit()
    button_pressed = press_t_2[0][0]
    time_button_pressed = str(globalClock.getTime())
    dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell}, time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


    #13. Display 300ms (18 frames) Cross Fixation Prior to Loop 2
    cross_type = "Pre_Loop_2_Cross"
    time_cross_shown = str(globalClock.getTime())
    for x in range(18):
        
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
        fixation_cross.draw()
        win.flip()
    time_cross_gone = str(globalClock.getTime())
    dataFile_intro_data.write(f"cross_type,{cross_type},time_cross_shown,{time_cross_shown},{empty_cell},time_cross_gone,{time_cross_gone}\n\n")

    #14. Loop 2

    Loop(seq_2_1, "seq_2_1", seq_2_2_ic, "seq_2_2_ic", seq_2_3, "seq_2_3")

    #15 Display Blank Infinite after Loop 2 (press b to continue)
    message_number = "After_Loop_2_Infinite_Blank"
    message_content = "Blank Screen (must press b to continue)"
    blank_infinite.draw()
    win.flip()
    time_message_shown = str(globalClock.getTime())
    press_b_2 = event.waitKeys(keyList=["b", 'escape'], timeStamped=True)
    if press_b_2[0][0] == 'escape':
        quitTime = str(globalClock.getTime())
        dataFile.write(f"\nQuit_Time,{quitTime}")
        dataFile.close()
        dataFile_intro_data.close()
        core.quit()
    button_pressed = press_b_2[0][0]
    time_button_pressed = str(globalClock.getTime())
    dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell}, time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


###########################################-Loop 3

def runLoop3():
    #16 Display Get Ready (press t to continue)
    message_number = "Pre_Loop_3_Get_Ready"
    message_content = "Get Ready"
    get_ready_message.draw()
    win.flip()
    time_message_shown = str(globalClock.getTime())
    press_t_3 = event.waitKeys(keyList=["t", 'escape'], timeStamped=True)
    if press_t_3[0][0] == 'escape':
        quitTime = str(globalClock.getTime())
        dataFile.write(f"\nQuit_Time,{quitTime}")
        dataFile.close()
        dataFile_intro_data.close()
        core.quit()
    button_pressed = press_t_3[0][0]
    time_button_pressed = str(globalClock.getTime())
    dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},button_pressed,{button_pressed},{empty_cell}, time_button_pressed,{time_button_pressed},{empty_cell}\n\n")


    #17. Display 300ms (18 frames) Cross Fixation Prior to Loop 3
    cross_type = "Pre_Loop_3_Cross"
    time_cross_shown = str(globalClock.getTime())
    for x in range(18):
        ##Quit Button During Code
        quitbutton = event.getKeys(keyList=['escape'])
        if len(quitbutton) > 0:
            for a in quitbutton:
                if a == 'escape':
                    quitTime = str(globalClock.getTime())
                    dataFile.write(f"\nQuit_Time,{quitTime}")
                    dataFile.close()
                    dataFile_intro_data.close()
                    core.quit()
        
        fixation_cross.draw()
        win.flip()
    time_cross_gone = str(globalClock.getTime())
    dataFile_intro_data.write(f"cross_type,{cross_type},time_cross_shown,{time_cross_shown},{empty_cell},time_cross_gone,{time_cross_gone}\n\n")

    #18 Loop 3
    Loop(seq_3_1, "seq_3_1", seq_3_2_ic, "seq_3_2_ic", seq_3_3, "seq_3_3")


if start_loop == 'loop_1':
    runLoop1()
    runLoop2()
    runLoop3()
elif start_loop == 'loop_2':
    runLoop2()
    runLoop3()
elif start_loop == 'loop_3':
    runLoop3()

#19 Display Goodbye Screen (for 20 seconds)
message_number = "Post_Loop_3_Goodbye_Message"
message_content = "Thanks for participating!"
time_message_shown = str(globalClock.getTime())
for x in range(1200):
    
    ##Quit Button During Code
    quitbutton = event.getKeys(keyList=['escape'])
    if len(quitbutton) > 0:
        for a in quitbutton:
            if a == 'escape':
                quitTime = str(globalClock.getTime())
                dataFile.write(f"\nQuit_Time,{quitTime}")
                dataFile.close()
                dataFile_intro_data.close()
                core.quit()
    
    goodbye_message.draw()
    win.flip()
time_message_gone = str(globalClock.getTime())
dataFile_intro_data.write(f"message_number,{message_number},{empty_cell},message_content,{message_content},{empty_cell},time_message_shown,{time_message_shown},{empty_cell},time_message_gone,{time_message_gone},{empty_cell}\n\n")

###-> Stopped here- will need to check difference between dataFile_intro and normal dataFile


#20. End
dataFile.close()
dataFile_intro_data.close()
print("Run completed")

































'''
image | duration | button press duration |

welcome_screen = 

for increment in sequence:
    run the fixation for certain amount of frames
    run increment for certain amount of frames
        display correct number of words on screen
        record key presses and other data
        record the frame at which buttons was pressed, convert that to milliseconds, and record
        
after increment in sequence, then implement planned pauses according to frames

'''


#the seqx_x need to have the number and answers blank for the personal words in case you want those words to appear a certain amount randomly
'''
Not necessary anymore
#get personal_word_order value from mainlist to determine if word order is set or random
random_or_set_order = mainlist[9]["personal_word_number"]
'''

'''
Not necessary anymore
#Do this if random_or_set_order is set to random. Will randomize the personal words dict list prior to insertion
if random_or_set_order == "random":
    l = len(personal_words)
    newlist = l * [0]
    
    for x in range(0,len(newlist)):
        rand_integer = random.randrange(0,l)
        newlist[x] = personal_words[rand_integer]
        personal_words.pop(rand_integer)
        l -= 1
        
    personal_words = newlist
'''

'''
Not necessary anymore
def personal_word_number_randomizer(preseqlist):
    for i in range(0,len(preseqlist)):
        if preseqlist[i]["type"] == "personal":
            randnumber = random.randrange(1,5)
            preseqlist[i]["number"] = randnumber
            preseqlist[i]["answer"] = f"{randnumber+1}"
'''

